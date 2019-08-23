import openpyxl

wb=openpyxl.load_workbook('C:\\NSE\\outputs\Masters.xlsx')
ws=wb['Sheet1']
code_list=ws['C2':'H12']



for i in code_list :
   for j in i:
       ws[j.coordinate].number_format = '#,##0.00'
#         print(j.coordinate, j.value)

#ws['D2'].number_format = '0.0%'

col_range = ws['C:F']
for col in col_range:
    ws.column_dimendions.width = 15

print('--- END OF ROW ---')

wb.save('C:\\NSE\\outputs\Masters.xlsx')


