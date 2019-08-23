# writing Change in OI for Calls into a file
import datetime as dt

import openpyxl
from openpyxl.workbook import Workbook


def create_input_data():

    dttime=dt.datetime.now().strftime("%H:%M:%S")
    readfile="C:\\NSE\\inputs\\NSEOptions.xlsx"
    writefile="C:\\NSE\\inputs\\NSEOptions1.xlsx"

    workbook_name = writefile
    wb = Workbook()
      
    wb.create_sheet('NSE')
    print(wb.sheetnames)
    std=wb['Sheet']
    wb.remove(std)
    wb.save(filename = writefile)
    print(wb.sheetnames)
    
    workbook_name = readfile
    wb1 = openpyxl.load_workbook(readfile, data_only=True)
    wb2 = openpyxl.load_workbook(writefile)
    sheet1 = wb1['Sheet2']
    sheet2 = wb2['NSE']
    max_rows=sheet1.max_row
    for row in sheet1['A1' :'U'+str(max_rows)]:
        
        for cell in row:
            sheet2[cell.coordinate].value = cell.value
            
            
    wb1.save(writefile)
