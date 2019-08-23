# writing Change in OI for Calls into a file
import datetime as dt

from openpyxl import load_workbook


def write_calls_oi():

    dttime=dt.datetime.now().strftime("%H:%M:%S")
    readfile="C:\\NSE\\inputs\\NSEOptions.xlsm"
    
    workbook_name = readfile

    ws1 = load_workbook(workbook_name)
    sheet = ws1["NSEOptions"]
    max_rows = sheet.max_row
        
    copydata1=[dttime,  "CALL"]
    

    for r in range(2,max_rows):
        if sheet['B'+str(r)].value is None :
               copydata1.append(0)
        else:
               copydata1.append(sheet['B'+str(r)].value)


    copydata2=[dttime,  "CALL"]
    
    for r in range(2,max_rows):
       if sheet['A'+str(r)].value is None :
          copydata2.append(0)
       else:
            copydata2.append(sheet['A'+str(r)].value)

#    print("copydata2 XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX")
#    print(copydata2)
    
    writefile='C:\\NSE\\outputs\\NSEOI-LOG-' + dt.datetime.now().strftime('%Y-%B-%d')+".xlsx"
    
    workbook_name = writefile

    ws2 = load_workbook(workbook_name)
    sheet = ws2["CHGOI"]
    sheet.append(copydata2)
    ws2.save(filename=workbook_name)
    # print("CHGOI Worksheet Appended")

    
    copydata3=[]
    delta1=[]
    
    sheet = ws2["CHGOI"]
    
    max_rows = sheet.max_row
    max_cols = sheet.max_column+1
    # print(max_rows, max_cols)

    if max_rows == 2:
        for c1 in range(3, max_cols+1):
            delta1= copydata2
            delta1[1]="DIFF-CALL-CHGOI"
    else:
        # print("XXXXXXXXXXXXXXXXXXXX in Else")
        max_rows = max_rows - 2
        # print(max_rows)
        for c1 in range(3,max_cols):
            copydata3.append(sheet.cell(column=c1,row=max_rows).value)
            
        del copydata2[0:2]
        delta1 = [copydata2 - copydata3 for copydata2, copydata3 in zip(copydata2,copydata3)]
        delta1.insert(0, dttime)
        delta1.insert(1, "DIFF-CALL-CHGOI")

    sheet = ws2["DIFFCHGOI"]
    
    workbook_name = writefile

    ws2 = load_workbook(workbook_name)
    sheet = ws2["DIFFCHGOI"]
    sheet.append(delta1)
    ws2.save(filename=workbook_name)
    # print("DIFFCHGOI Worksheet Appended")

    workbook_name = readfile

    ws1 = load_workbook(workbook_name)
    sheet = ws1["NSEOptions"]
    max_rows = sheet.max_row
            
    copydata1=[dttime,  "CALL"]

    for r in range(2,max_rows):
        if sheet['B'+str(r)].value is None :
               copydata1.append(0)
        else:
               copydata1.append(sheet['B'+str(r)].value)

        
#    print(copydata1)

    copydata2=[dttime,  "CALL"]
    
    for r in range(2,max_rows):
       if sheet['A'+str(r)].value is None :
          copydata2.append(0)
       else:
            copydata2.append(sheet['A'+str(r)].value)

    workbook_name = writefile
    ws2 = load_workbook(workbook_name)
    sheet = ws2["OI"]
    sheet.append(copydata1)
    ws2.save(filename=workbook_name)
    
    print("Writing Call Details Completed ...")
    
       
    
    
