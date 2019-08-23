# Writing Strike Price at the time of file creations
import datetime as dt

from openpyxl import load_workbook


def write_strike_price():

    readfile="C:\\NSE\\inputs\\NSEOptions.xlsm"
    dttime=dt.datetime.now().strftime("%Y-%b-%d %H:%M:%S")
    writefile='C:\\NSE\\outputs\\NSEOI-LOG-' + dt.datetime.now().strftime('%Y-%B-%d')+".xlsx"

    workbook_name = readfile
    print(readfile)

    ws1 = load_workbook(workbook_name)
    sheet = ws1["NSEOptions"]
        
    max_rows = sheet.max_row
    copydata=[dttime, 'Strike Price']
 
    for r in range(2,max_rows):
        copydata.append(sheet['K'+str(r)].value)

    # print(copydata)

    workbook_name = writefile
    print(workbook_name)

    ws2 = load_workbook(workbook_name)
    sheet = ws2["CHGOI"]
    sheet.append(copydata)
    ws2.save(filename=workbook_name)
    sheet = ws2["OI"]
    sheet.append(copydata)
    sheet = ws2["DIFFCHGOI"]
    sheet.append(copydata)
    sheet = ws2["DIFFOI"]
    sheet.append(copydata)
    ws2.save(filename=workbook_name)
