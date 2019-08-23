# writing Change in OI for Puts into a file
import datetime as dt

from openpyxl import load_workbook


def write_put_oi():

    dttime=dt.datetime.now().strftime("%H:%M:%S")
    readfile="C:\\NSE\\inputs\\NSEOptions.xlsm"
    writefile='C:\\NSE\\outputs\\NSEOI-LOG-' + dt.datetime.now().strftime('%Y-%B-%d')+".xlsx"
    
    workbook_name = readfile

    ws1 = load_workbook(workbook_name)
    sheet = ws1["NSEOptions"]
    max_rows = sheet.max_row
        
    copydata1=[dttime,  "PUT"]

    for r in range(2,max_rows):
        if sheet['T'+str(r)].value is None :
               copydata1.append(0)
        else:
               copydata1.append(sheet['T'+str(r)].value)

    workbook_name = writefile
    ws2 = load_workbook(workbook_name)
    sheet = ws2["CHGOI"]
    sheet.append(copydata1)
    ws2.save(filename=workbook_name)

    copydata3=[]
    delta1=[]
    
    sheet = ws2["CHGOI"]
    max_rows = sheet.max_row
    max_cols = sheet.max_column+1
    # print("Max Rows : " +str(max_rows))
    # print("Max Cols : " +str(max_cols))

    if max_rows == 3 :
       delta1= copydata1
       delta1[1]="DIFF-PUT-CHGOI"
    else:
        max_rows = max_rows - 2
        # print("Max Rows in Else : " +str(max_rows))
        # print(sheet.cell(column=3,row=3).value)
        # print(sheet.cell(column=4,row=4).value)
        
        # print("----------------")
        for c1 in range(3,max_cols):
            # print(c1)
            # print(sheet.cell(column=c1,row=max_rows).value)
            copydata3.append(sheet.cell(column=c1,row=max_rows).value)

                   
        del copydata1[0:2]
        # print(copydata1)
        # print(copydata3)
        delta1 = [copydata1 - copydata3 for copydata1, copydata3 in zip(copydata1,copydata3)]
        delta1.insert(0, dttime)
        delta1.insert(1, "DIFF-PUT-CHGOI")

    sheet = ws2["DIFFCHGOI"]
    
    workbook_name = writefile

    ws2 = load_workbook(workbook_name)
    sheet = ws2["DIFFCHGOI"]
    sheet.append(delta1)
    ws2.save(filename=workbook_name)
    # print("DIFFCHGOI Worksheet Appended")
 
        
#    print(copydata1)

    workbook_name = readfile
    ws1 = load_workbook(workbook_name)
    sheet = ws1["NSEOptions"]
    max_rows = sheet.max_row
    max_cols = sheet.max_column+1

    copydata2=[dttime,  "PUT"]
    
    for r in range(2,max_rows):
       if sheet['U'+str(r)].value is None :
          copydata2.append(0)
       else:
            copydata2.append(sheet['U'+str(r)].value)

#    print(copydata2)
    
    copydata=[dttime,  "PUT"]
    workbook_name = writefile
    ws2 = load_workbook(workbook_name)
    sheet = ws2["OI"]
    sheet.append(copydata2)
    ws2.save(filename=workbook_name)

    print("Writing Put Details Completed ...")

