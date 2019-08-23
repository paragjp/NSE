# writing Change in OI for Puts into a file
import datetime as dt

from openpyxl import load_workbook


def write_put_oi():

    dttime=dt.datetime.now().strftime("%H:%M:%S")
    readfile="C:\\NSE\\inputs\\NSEOptions.xlsm"
    
    workbook_name = readfile

    ws1 = load_workbook(workbook_name)
    sheet = ws1["NSEOptions"]
    max_rows = sheet.max_row
        
    copydata1=[dttime,  "PUT"]

    for r in range(2,max_rows):
        if sheet['T'+str(r)].value is None :
               copydata1.append(0)
        else:
               copydata1.append(sheet['T'+str(r)].value)

        
    print(copydata1)

    copydata2=[dttime,  "PUT"]
    
    for r in range(2,max_rows):
       if sheet['U'+str(r)].value is None :
          copydata2.append(0)
       else:
            copydata2.append(sheet['U'+str(r)].value)

    print(copydata2)
    
    writefile='C:\\NSE\\outputs\\NSEOI-LOG-' + dt.datetime.now().strftime('%Y-%B-%d')+".xlsx"
    
    workbook_name = writefile
    ws2 = load_workbook(workbook_name)
    sheet = ws2["CHGOI"]
    sheet.append(copydata1)
    ws2.save(filename=workbook_name)


    copydata=[dttime,  "PUT"]
    workbook_name = writefile
    ws2 = load_workbook(workbook_name)
    sheet = ws2["OI"]
   
    sheet.append(copydata2)
    ws2.save(filename=workbook_name)

