# create OI Changes Recording Log file

import datetime as dt
import os

from openpyxl.workbook import Workbook
from write_calls_oi import write_calls_oi
from write_put_oi import write_put_oi
from write_strike_price import write_strike_price

dttime=dt.datetime.now().strftime("%d-%b-%Y %H:%M:%S")
heading=[" Changes in OI for Options as on " + dttime[:11]]
filename='C:\\NSE\\outputs\\NSEOI-LOG-' + dt.datetime.now().strftime('%Y-%B-%d')+".xlsx"

# print(os.path.isfile(filename))

if os.path.isfile(filename):
  os.remove(filename)
  print("File Removed")

data=dttime
workbook_name = filename
wb = Workbook()
  
wb.create_sheet('CHGOI')
wb.create_sheet('OI')
wb.create_sheet('DIFFCHGOI')
wb.create_sheet('DIFFOI')

# print(wb.sheetnames)
  
std=wb['Sheet']
wb.remove(std)
wb.save(filename = workbook_name)
# print(wb.sheetnames)
  
print("File Created")
write_strike_price()
write_calls_oi()
write_put_oi()
